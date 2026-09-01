# -*- coding: utf-8 -*-
"""
analysis_core.py
------------------
純運算邏輯層：讀取Excel、篩選範圍、繪圖、統計分析(Cpk / 相關性 / 敘述統計)、
以及匯出完整Excel報表。刻意與GUI (tkinter) 分離，方便獨立測試與重複使用。
"""
import io
import os
from datetime import datetime

import numpy as np
import pandas as pd
from scipy import stats

import matplotlib
matplotlib.use("Agg")  # 供無畫面環境/匯出圖片使用；GUI程式會另外用 TkAgg 畫布
import matplotlib.pyplot as plt
from matplotlib import font_manager


def _setup_cjk_font():
    """自動偵測系統中可顯示中文的字型，避免圖表中文字變成方框。
    在不同作業系統(Windows/macOS/Linux)上都會嘗試常見字型名稱。"""
    candidates = [
        "Microsoft JhengHei", "Microsoft YaHei", "PingFang TC", "PingFang SC",
        "Noto Sans CJK TC", "Noto Sans CJK SC", "Noto Sans TC", "Noto Sans SC",
        "Heiti TC", "SimHei", "WenQuanYi Zen Hei", "Arial Unicode MS",
    ]
    available = {f.name for f in font_manager.fontManager.ttflist}
    for name in candidates:
        if name in available:
            matplotlib.rcParams["font.sans-serif"] = [name]
            matplotlib.rcParams["font.family"] = "sans-serif"
            matplotlib.rcParams["axes.unicode_minus"] = False
            return name
    matplotlib.rcParams["axes.unicode_minus"] = False
    return None


_ACTIVE_CJK_FONT = _setup_cjk_font()

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

CHART_TYPES = ["長條圖", "折線圖", "圓餅圖", "直方圖", "散佈圖", "盒鬚圖"]
ANALYSIS_METHODS = ["Cpk 製程能力分析", "相關性分析", "敘述統計"]


# ---------------------------------------------------------------------------
# 1. 匯入 Excel
# ---------------------------------------------------------------------------
def list_sheet_names(filepath):
    """回傳Excel檔案中所有分頁名稱"""
    xls = pd.ExcelFile(filepath)
    return xls.sheet_names


def load_sheet(filepath, sheet_name):
    """讀取指定分頁為 DataFrame"""
    df = pd.read_excel(filepath, sheet_name=sheet_name)
    return df


# ---------------------------------------------------------------------------
# 2. 依欄(column)或行(row)選擇範圍
# ---------------------------------------------------------------------------
def filter_range(df, columns=None, row_start=None, row_end=None):
    """
    columns: 欲保留的欄位名稱清單 (None = 全部)
    row_start / row_end: 資料列範圍 (1-based，對應資料本身，不含表頭)，None = 不限制
    """
    result = df.copy()
    if columns:
        result = result[[c for c in columns if c in result.columns]]
    if row_start is not None or row_end is not None:
        start_idx = (row_start - 1) if row_start else 0
        end_idx = row_end if row_end else len(result)
        start_idx = max(start_idx, 0)
        end_idx = min(end_idx, len(result))
        result = result.iloc[start_idx:end_idx]
    return result.reset_index(drop=True)


def numeric_columns(df):
    return [c for c in df.columns if pd.api.types.is_numeric_dtype(df[c])]


# ---------------------------------------------------------------------------
# 3. 圖表繪製 (含 USL / LSL 上下限)
# ---------------------------------------------------------------------------
def draw_chart(fig, df, columns, chart_type, spec_limits=None, x_col=None, y_col=None, title=None):
    """
    在傳入的 matplotlib Figure 上繪製圖表。
    spec_limits: dict{col: {"LSL": float or None, "USL": float or None}}
    x_col/y_col: 散佈圖使用
    回傳: 使用的 Axes (方便呼叫端進一步調整)
    """
    fig.clf()
    ax = fig.add_subplot(111)
    spec_limits = spec_limits or {}

    if chart_type == "長條圖":
        df[columns].mean(numeric_only=True).plot(kind="bar", ax=ax, color="#4C72B0")
        _draw_spec_lines(ax, columns, spec_limits, horizontal=True)
        ax.set_ylabel("平均值")

    elif chart_type == "折線圖":
        for col in columns:
            ax.plot(df.index, df[col], marker="o", markersize=3, label=col)
        _draw_spec_lines(ax, columns, spec_limits, horizontal=True)
        ax.set_xlabel("資料筆數 (index)")
        ax.legend(loc="best", fontsize=8)

    elif chart_type == "圓餅圖":
        col = columns[0]
        series = df[col]
        if pd.api.types.is_numeric_dtype(series):
            # 數值型：以區間分組計數
            counts = pd.cut(series.dropna(), bins=5).value_counts().sort_index()
            counts.index = [str(i) for i in counts.index]
        else:
            counts = series.value_counts()
        ax.pie(counts.values, labels=counts.index, autopct="%1.1f%%", startangle=90)
        ax.set_title(f"{col} 分布比例")

    elif chart_type == "直方圖":
        for col in columns:
            ax.hist(df[col].dropna(), bins=20, alpha=0.6, label=col)
        _draw_spec_lines(ax, columns, spec_limits, horizontal=False)
        ax.set_xlabel("數值")
        ax.set_ylabel("次數")
        ax.legend(loc="best", fontsize=8)

    elif chart_type == "散佈圖":
        x_col = x_col or columns[0]
        y_col = y_col or (columns[1] if len(columns) > 1 else columns[0])
        ax.scatter(df[x_col], df[y_col], alpha=0.7, color="#55A868")
        ax.set_xlabel(x_col)
        ax.set_ylabel(y_col)
        _draw_spec_lines(ax, [y_col], spec_limits, horizontal=True)

    elif chart_type == "盒鬚圖":
        data = [df[col].dropna() for col in columns]
        ax.boxplot(data, labels=columns, showmeans=True)
        _draw_spec_lines(ax, columns, spec_limits, horizontal=True)

    else:
        raise ValueError(f"不支援的圖表類型: {chart_type}")

    ax.set_title(title or chart_type)
    fig.tight_layout()
    return ax


def _draw_spec_lines(ax, columns, spec_limits, horizontal=True):
    """在圖上畫出 USL/LSL 參考線 (取第一個有設定規格的欄位)"""
    for col in columns:
        limits = spec_limits.get(col)
        if not limits:
            continue
        usl, lsl = limits.get("USL"), limits.get("LSL")
        if usl is not None:
            if horizontal:
                ax.axhline(usl, color="red", linestyle="--", linewidth=1, label=f"{col} USL={usl}")
            else:
                ax.axvline(usl, color="red", linestyle="--", linewidth=1, label=f"{col} USL={usl}")
        if lsl is not None:
            if horizontal:
                ax.axhline(lsl, color="orange", linestyle="--", linewidth=1, label=f"{col} LSL={lsl}")
            else:
                ax.axvline(lsl, color="orange", linestyle="--", linewidth=1, label=f"{col} LSL={lsl}")


def fig_to_png_bytes(fig, dpi=150):
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=dpi, bbox_inches="tight")
    buf.seek(0)
    return buf


# ---------------------------------------------------------------------------
# 4. 統計分析
# ---------------------------------------------------------------------------
def calculate_cpk(series, lsl=None, usl=None):
    """計算單一欄位的 Cp / Cpk / Cpu / Cpl"""
    data = pd.to_numeric(series, errors="coerce").dropna()
    n = len(data)
    mean = float(data.mean()) if n else np.nan
    std = float(data.std(ddof=1)) if n > 1 else np.nan

    cpu = cpl = cpk = cp = None
    if std and std > 0 and not np.isnan(std):
        if usl is not None:
            cpu = (usl - mean) / (3 * std)
        if lsl is not None:
            cpl = (mean - lsl) / (3 * std)
        if usl is not None and lsl is not None:
            cp = (usl - lsl) / (6 * std)
            cpk = min(cpu, cpl)
        elif usl is not None:
            cpk = cpu
        elif lsl is not None:
            cpk = cpl

    def _round(v):
        return round(v, 4) if isinstance(v, (int, float)) and not np.isnan(v) else v

    return {
        "n": n,
        "平均值": _round(mean),
        "標準差": _round(std),
        "LSL": lsl,
        "USL": usl,
        "Cp": _round(cp) if cp is not None else None,
        "Cpu": _round(cpu) if cpu is not None else None,
        "Cpl": _round(cpl) if cpl is not None else None,
        "Cpk": _round(cpk) if cpk is not None else None,
        "製程評估": _evaluate_cpk(cpk),
    }


def _evaluate_cpk(cpk):
    if cpk is None or (isinstance(cpk, float) and np.isnan(cpk)):
        return "未設定規格"
    if cpk >= 1.33:
        return "能力充分"
    if cpk >= 1.0:
        return "能力尚可"
    return "能力不足"


def calculate_cpk_batch(df, columns, spec_limits):
    """回傳多欄位的 Cpk 分析結果 DataFrame"""
    rows = []
    for col in columns:
        limits = spec_limits.get(col, {}) if spec_limits else {}
        result = calculate_cpk(df[col], lsl=limits.get("LSL"), usl=limits.get("USL"))
        result_row = {"欄位": col}
        result_row.update(result)
        rows.append(result_row)
    return pd.DataFrame(rows)


def calculate_correlation(df, columns, method="pearson"):
    """回傳相關係數矩陣 DataFrame"""
    valid_cols = [c for c in columns if pd.api.types.is_numeric_dtype(df[c])]
    corr = df[valid_cols].corr(method=method)
    return corr


def calculate_descriptive(df, columns):
    """敘述統計：count/mean/std/min/25%/50%/75%/max + skew/kurtosis"""
    valid_cols = [c for c in columns if pd.api.types.is_numeric_dtype(df[c])]
    desc = df[valid_cols].describe().T
    desc["偏態(skew)"] = df[valid_cols].skew()
    desc["峰度(kurtosis)"] = df[valid_cols].kurtosis()
    desc = desc.round(4)
    desc.index.name = "欄位"
    return desc.reset_index()


# ---------------------------------------------------------------------------
# 5. 匯出完整 Excel 報表
# ---------------------------------------------------------------------------
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def _write_dataframe(ws, df, start_row=1):
    for j, col in enumerate(df.columns, start=1):
        cell = ws.cell(row=start_row, column=j, value=str(col))
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    for i, (_, row) in enumerate(df.iterrows(), start=start_row + 1):
        for j, val in enumerate(row, start=1):
            if isinstance(val, (np.integer,)):
                val = int(val)
            elif isinstance(val, (np.floating,)):
                val = float(val)
            elif pd.isna(val):
                val = None
            ws.cell(row=i, column=j, value=val)
    for j, col in enumerate(df.columns, start=1):
        max_len = max([len(str(col))] + [len(str(v)) for v in df[col].astype(str).values[:200]])
        ws.column_dimensions[get_column_letter(j)].width = min(max(max_len + 2, 10), 40)
    return start_row + len(df) + 1


def export_report(output_path, source_df, charts=None, analyses=None, spec_limits=None,
                   sheet_prefix="資料"):
    """
    output_path: 輸出檔案路徑
    source_df: 篩選後的原始資料 DataFrame
    charts: list of dict {"title": str, "png_bytes": BytesIO}
    analyses: list of dict {"title": str, "df": DataFrame}
    spec_limits: dict{col: {"LSL":..,"USL":..}}
    """
    charts = charts or []
    analyses = analyses or []
    spec_limits = spec_limits or {}

    wb = Workbook()

    # -- 分頁1：原始資料 --
    ws_data = wb.active
    ws_data.title = f"{sheet_prefix}"
    _write_dataframe(ws_data, source_df, start_row=1)

    # -- 分頁2：規格上下限 --
    if spec_limits:
        ws_spec = wb.create_sheet("規格上下限")
        spec_df = pd.DataFrame([
            {"欄位": col, "LSL": v.get("LSL"), "USL": v.get("USL")}
            for col, v in spec_limits.items()
        ])
        _write_dataframe(ws_spec, spec_df, start_row=1)

    # -- 分頁3：圖表 --
    if charts:
        ws_chart = wb.create_sheet("圖表")
        row_cursor = 1
        for item in charts:
            ws_chart.cell(row=row_cursor, column=1, value=item["title"]).font = Font(bold=True, size=12)
            row_cursor += 1
            img = XLImage(item["png_bytes"])
            img.width, img.height = 640, 420
            anchor = f"A{row_cursor}"
            ws_chart.add_image(img, anchor)
            row_cursor += 24  # 預留圖片高度空間

    # -- 分頁4起：分析結果 (每個分析一個分頁) --
    used_names = set()
    for item in analyses:
        base_name = item["title"][:25] or "分析結果"
        name = base_name
        n = 1
        while name in used_names or name in wb.sheetnames:
            n += 1
            name = f"{base_name}_{n}"
        used_names.add(name)
        ws = wb.create_sheet(name)
        _write_dataframe(ws, item["df"], start_row=1)

    # -- 匯出資訊 --
    ws_info = wb.create_sheet("報表資訊", 0)
    ws_info["A1"] = "Excel 數據分析報表"
    ws_info["A1"].font = Font(bold=True, size=14)
    ws_info["A3"] = "產生時間"
    ws_info["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws_info["A4"] = "資料筆數"
    ws_info["B4"] = len(source_df)
    ws_info["A5"] = "欄位數"
    ws_info["B5"] = len(source_df.columns)
    ws_info.column_dimensions["A"].width = 16
    ws_info.column_dimensions["B"].width = 24

    wb.save(output_path)
    return output_path
